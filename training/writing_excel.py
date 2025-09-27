'''
openpyxl    :   To write into the excel, we use openpyxl

To install openpyxl
    Go to command prompt    --> pip install openpyxl
'''

# from openpyxl import Workbook
#
# ## create the excel workbook
# workbook = Workbook()
#
# ## Initialize worksheet
# worksheet = workbook.active
#
# ## setup the sheetname(optional)
# worksheet.title = 'candidates_info'
#
# ## Enter the data
# worksheet['A1'] = 'name'
# worksheet['B1'] = 'place'
# worksheet['C1'] = 'email_id'
# worksheet['D1'] = 'phone_num'
#
# data_list = [
#     ['Shailaja', 'Bengaluru', 'shailaja@gmail.com', '9080706050'],
#     ['Deeksha', 'Chennai', 'deeksha@gmail.com', '9181716151'],
#     ['Hashmath', 'Hyderabad', 'hashmath@gmail.com', '9282726252'],
#     ['Atharva', 'Pune', 'atharva@gmail.com', '9876987655']
# ]
# for data in data_list:
#     worksheet.append(data)
#
# # ## save the excel file
# # workbook.save('M3_candidates_data.xlsx')
#
# ## To save the excel file in different location
# workbook.save(r'C:\Users\Ramya\PycharmProjects\sel-M3-weekend-Aug16-2025\files_\M3_data.xlsx')

####################################################################################

## ANALYZE THE BELOW CODE

from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment

## create the excel workbook
excel_workbook = Workbook()

## initialize the worksheet
worksheet = excel_workbook.active

## set a name for the worksheet(optional)
worksheet.title = 'personal_information'

## Adding the data
worksheet['A1'] = 'name'
worksheet['B1'] = 'place'
worksheet['C1'] = 'salary'
worksheet['D1'] = 'email_id'

## Font formatting
worksheet['A1'].font = Font(name='Arial', bold=True, italic=True, color='FF0000')   ## formats individual cell


data_list = [
    ['Vihaan', 'Delhi', 50000, 'vihaan@gmail.com'],
    ['Siddharth', 'Mumbai', 55000, 'siddu@gmail.com'],
    ['Saina', 'Hyderabad', 60000, 'saina@gmail.com'],
    ['Risto', 'Pune', 65000, 'risto@gmail.com'],
    ['Goku', 'Assam', 70000, 'goku@gmail.com'],
    ['Paresh', 'Gujrat', 75000, 'paresh@gmail.com']
]

for row in data_list:
    worksheet.append(row)

# ## save the excel file
# excel_workbook.save('emp_data.xlsx')

##-----------------------------------------------------------------------------
# ## formatting entire row
# ## define styles
# bold_font = Font(bold=True, color='FFFFFF')
# fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
# for cell in worksheet[1]:
#     cell.font = bold_font
#     cell.fill = fill
#
# ## save the excel file
# excel_workbook.save('emp_data.xlsx')

# ##-----------------------------------------------------------------------------

## Add new sheet to the existing workbook

## load existing workbook
workbook = load_workbook('emp_data.xlsx')

## create a new sheet
new_sheet = workbook.create_sheet(title="title")

## write the data asusal
## save the file

# ##-----------------------------------------------------------------------------
#
# ## save the excel file
# excel_workbook.save('emp_data.xlsx')
#
# # ## save the excel file in different location
# # excel_workbook.save(r'C:\Users\Ramya\PycharmProjects\Sel-M6-April1\files_\emp_data.xlsx')
#
# ##-----------------------------------------------------------------------------
#
# ## Clear the cell's value
#
# from openpyxl import load_workbook
#
# workbook = load_workbook("emp_data.xlsx")
# worksheet = workbook.active
#
# worksheet['A1'] = None      ## Clears the value and formats in cell A1
#
# workbook.save("emp_data.xlsx")
#
# ##-----------------------------------------------------------------------------
#
# ## Delete entire row or column
#
# from openpyxl import load_workbook
#
# workbook = load_workbook("emp_data.xlsx")
# worksheet = workbook.active
#
# worksheet.delete_rows(1)        ## Deletes row1
# worksheet.delete_cols(1)        ## Deletes col1
















































































